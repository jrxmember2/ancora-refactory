import sys
import json
import os
from datetime import datetime

HEADER_ALIASES = {
    'CONDOMINIO', 'NOMECONDOMINIO',
    'BLOCO', 'TORRE',
    'UNIDADE', 'UNID',
    'REFERENCIA', 'COMPETENCIA', 'MESREF', 'MES', 'MESREF',
    'VENCIMENTO', 'DATAVENCIMENTO',
    'VALOR', 'TOTAL', 'VALORATUALIZADO', 'VALORORIGINAL',
}

REQUIRED_GROUPS = {
    'condominium': {'CONDOMINIO', 'NOMECONDOMINIO'},
    'unit': {'UNIDADE', 'UNID'},
    'reference': {'REFERENCIA', 'COMPETENCIA', 'MESREF', 'MES'},
    'due_date': {'VENCIMENTO', 'DATAVENCIMENTO'},
    'amount': {'VALOR', 'TOTAL', 'VALORATUALIZADO', 'VALORORIGINAL'},
}


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def normalize_header(value):
    text = stringify(value).strip().upper()
    replacements = str.maketrans({
        'Á': 'A', 'À': 'A', 'Â': 'A', 'Ã': 'A', 'Ä': 'A',
        'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
        'Í': 'I', 'Ì': 'I', 'Î': 'I', 'Ï': 'I',
        'Ó': 'O', 'Ò': 'O', 'Ô': 'O', 'Õ': 'O', 'Ö': 'O',
        'Ú': 'U', 'Ù': 'U', 'Û': 'U', 'Ü': 'U',
        'Ç': 'C',
    })
    text = text.translate(replacements)
    return ''.join(ch for ch in text if ch.isalnum())


def load_xlsx_sheets(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([stringify(cell) for cell in row])
        sheets.append((ws.title, rows))
    return sheets


def load_xls_sheets(path):
    import xlrd

    book = xlrd.open_workbook(path)
    sheets = []
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            current = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(value, book.datemode)
                        value = dt.date().isoformat()
                    except Exception:
                        pass
                current.append(stringify(value))
            rows.append(current)
        sheets.append((sheet.name, rows))
    return sheets


def header_groups(row):
    normalized = [normalize_header(v) for v in row]
    found = set()
    for item in normalized:
        for group, aliases in REQUIRED_GROUPS.items():
            if item in aliases:
                found.add(group)
    return found, normalized


def detect_headers(rows):
    best_idx = None
    best_score = -1
    best_row = []

    for idx, row in enumerate(rows[:20]):
        found, normalized = header_groups(row)
        score = len(found)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_row = row
        if score >= 5:
            return idx, row, score

    if best_idx is not None and best_score >= 4:
        return best_idx, best_row, best_score

    return 0, (rows[0] if rows else []), 0


def select_sheet(sheets):
    best = None
    best_score = -1
    for sheet_name, rows in sheets:
        header_idx, header, score = detect_headers(rows)
        if score > best_score:
            best = (sheet_name, rows, header_idx, header, score)
            best_score = score
    return best


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Arquivo não informado."}))
        sys.exit(1)

    path = sys.argv[1]
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == '.xlsx':
            sheets = load_xlsx_sheets(path)
        elif ext == '.xls':
            sheets = load_xls_sheets(path)
        else:
            raise ValueError('Extensão não suportada')
    except Exception as exc:
        print(json.dumps({"error": str(exc)}))
        sys.exit(2)

    selected = select_sheet(sheets)
    if not selected:
        print(json.dumps({"error": "A planilha não contém abas legíveis."}, ensure_ascii=False))
        sys.exit(3)

    sheet_name, rows, header_idx, header, score = selected
    if score < 4:
        print(json.dumps({
            "error": "Cabeçalhos obrigatórios não encontrados. Use: Condomínio, Bloco (opcional), Unidade, Referência, Vencimento e Valor."
        }, ensure_ascii=False))
        sys.exit(4)

    data_rows = rows[header_idx + 1:] if rows else []

    print(json.dumps({
        "sheet_name": sheet_name,
        "header_row_index": header_idx + 1,
        "headers": [stringify(item) for item in header],
        "rows": data_rows,
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
