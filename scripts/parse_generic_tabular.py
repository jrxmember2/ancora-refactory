import csv
import json
import os
import sys


def read_csv(path):
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as fh:
                rows = list(csv.reader(fh, delimiter=";"))
                if rows:
                    return rows
        except Exception:
            continue

    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as fh:
        return list(csv.reader(fh, delimiter=","))


def read_xlsx(path):
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value).strip() for value in row])
    return rows


def main():
    if len(sys.argv) < 2:
        raise RuntimeError("Arquivo nao informado.")

    path = sys.argv[1]
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        rows = read_csv(path)
    elif ext == ".xlsx":
        rows = read_xlsx(path)
    else:
        raise RuntimeError("Formato nao suportado. Use CSV ou XLSX.")

    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    print(json.dumps({
        "headers": headers,
        "rows": data_rows,
    }, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        sys.exit(1)
